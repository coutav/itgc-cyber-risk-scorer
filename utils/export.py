"""
utils/export.py
Generates Aura-compatible JSON export and Excel batch output.
"""

import json
import io
from datetime import datetime
import pandas as pd


def to_aura_json(result: dict, metadata: dict) -> str:
    """
    Produce an Aura-compatible JSON string from a single scoring result.

    Parameters
    ----------
    result   : output dict from predict_risk()
    metadata : dict with keys: observation, risk, control_domain,
               application, industry, app_type, auditor_override,
               override_reason, override_by
    """
    payload = {
        "schema_version": "1.0",
        "tool": "ITGC Cyber Risk Scoring Model",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "finding": {
            "control_domain":   metadata.get("control_domain", ""),
            "application":      metadata.get("application", ""),
            "industry":         metadata.get("industry", ""),
            "app_type":         metadata.get("app_type", ""),
            "observation_text": metadata.get("observation", "")[:500],
            "risk_text":        metadata.get("risk", "")[:500],
        },
        "model_output": {
            "risk_score":       result.get("risk_score"),
            "risk_band":        result.get("risk_band"),
            "predicted_class":  result.get("predicted_class"),
            "probabilities": {
                "p_low":    result.get("p_low"),
                "p_medium": result.get("p_medium"),
                "p_high":   result.get("p_high"),
            },
            "cyber_risk_flags": result.get("flags", {}),
        },
        "explainability": {
            "method": "SHAP (TreeExplainer)",
            "top_features": [
                {"feature": f, "shap_value": round(v, 4)}
                for f, v in zip(
                    result.get("shap_features", [])[:10],
                    result.get("shap_values",   [])[:10],
                )
            ],
        },
        "auditor_review": {
            "model_accepted":    not metadata.get("auditor_override", False),
            "override_applied":  metadata.get("auditor_override", False),
            "override_band":     metadata.get("override_band", None),
            "override_reason":   metadata.get("override_reason", None),
            "reviewed_by":       metadata.get("override_by", None),
            "reviewed_at":       datetime.utcnow().isoformat() + "Z"
                                 if metadata.get("auditor_override") else None,
        },
    }
    return json.dumps(payload, indent=2)


def batch_to_excel(df: pd.DataFrame) -> bytes:
    """Convert batch results DataFrame to a formatted Excel workbook in memory."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Risk Scores")

        wb = writer.book
        ws = writer.sheets["Risk Scores"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header style
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="06B6D4", size=10)
        thin = Side(border_style="thin", color="1E2D45")
        border = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border

        # Band colour rows
        band_fills = {
            "High":   PatternFill("solid", fgColor="3B0000"),
            "Medium": PatternFill("solid", fgColor="2D1A00"),
            "Low":    PatternFill("solid", fgColor="002D1A"),
        }
        band_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Risk Band":
                band_col = idx
                break

        if band_col:
            for row in ws.iter_rows(min_row=2):
                band = row[band_col - 1].value
                if band in band_fills:
                    for cell in row:
                        cell.fill = band_fills[band]

        # Auto-width
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Freeze header
        ws.freeze_panes = "A2"

    return buf.getvalue()
